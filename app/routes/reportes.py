"""
Módulo de Reportes - Sistema de Gestión Biblioteca/Almacén SENA
Genera reportes de inventario, préstamos y usuarios con exportación Excel/PDF.
Respeta estrictamente el RBAC de 5 roles.
"""
from flask import (
    Blueprint, render_template, request, flash, redirect,
    url_for, send_file, current_app, abort
)
from flask_login import login_required, current_user
from app import db
from app.models.equipos import Equipo
from app.models.libros import Libro
from app.models.prestamos import Prestamo
from app.models.prestamos_libros import PrestamoLibro
from app.models.usuarios import Usuario
from datetime import datetime, timezone, timedelta
import io

bp = Blueprint('reportes', __name__, url_prefix='/reportes')

# Helper functions para control de RBAC local
def _puede_ver_equipos():
    return current_user.rol in ['administrador', 'almacenista']

def _puede_ver_libros():
    return current_user.rol in ['administrador', 'bibliotecario']

def _es_usuario_basico():
    return current_user.rol in ['aprendiz', 'instructor']

def _parse_fecha(fecha_str):
    if not fecha_str:
        return None
    try:
        return datetime.strptime(fecha_str, '%Y-%m-%d')
    except ValueError:
        return None


# ─── ÍNDICE DE REPORTES ───────────────────────────────────────
@bp.route('/')
@login_required
def index():
    return render_template('reportes/index.html')


# ─── INVENTARIO ───────────────────────────────────────────────
@bp.route('/inventario')
@login_required
def inventario():
    if _es_usuario_basico():
        flash('No tienes permisos para acceder a reportes de inventario.', 'danger')
        return redirect(url_for('reportes.index'))

    estado = request.args.get('estado', '')
    tipo = request.args.get('tipo', '')
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')

    equipos_data = []
    libros_data = []
    tipos_equipo = []

    if _puede_ver_equipos():
        q = Equipo.query.filter_by(eliminado=False)
        if estado:
            q = q.filter(Equipo.estado == estado)
        if tipo:
            q = q.filter(Equipo.tipo_equipo == tipo)
        if _parse_fecha(fecha_inicio):
            q = q.filter(Equipo.fecha_registro >= _parse_fecha(fecha_inicio))
        if _parse_fecha(fecha_fin):
            q = q.filter(Equipo.fecha_registro <= _parse_fecha(fecha_fin) + timedelta(days=1))
        equipos_data = q.order_by(Equipo.fecha_registro.desc()).all()
        tipos_equipo = db.session.query(Equipo.tipo_equipo).filter_by(eliminado=False).distinct().all()

    if _puede_ver_libros():
        q = Libro.query.filter_by(eliminado=False)
        if estado:
            q = q.filter(Libro.estado == estado)
        if _parse_fecha(fecha_inicio):
            q = q.filter(Libro.fecha_registro >= _parse_fecha(fecha_inicio))
        if _parse_fecha(fecha_fin):
            q = q.filter(Libro.fecha_registro <= _parse_fecha(fecha_fin) + timedelta(days=1))
        libros_data = q.order_by(Libro.fecha_registro.desc()).all()

    return render_template('reportes/inventario.html',
        equipos=equipos_data, libros=libros_data,
        tipos_equipo=tipos_equipo,
        estado=estado, tipo=tipo,
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin
    )


# ─── PRÉSTAMOS GLOBALES ──────────────────────────────────────
@bp.route('/prestamos')
@login_required
def prestamos():
    if _es_usuario_basico():
        flash('No tienes permisos para acceder a reportes de préstamos.', 'danger')
        return redirect(url_for('reportes.index'))

    estado = request.args.get('estado', '')
    tipo_recurso = request.args.get('tipo_recurso', '')
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')

    prestamos_equipos = []
    prestamos_libros = []

    if _puede_ver_equipos() and tipo_recurso != 'libros':
        q = Prestamo.query.join(Equipo).join(Usuario, Prestamo.id_usuario == Usuario.id_usuario)
        if estado:
            q = q.filter(Prestamo.estado == estado)
        if _parse_fecha(fecha_inicio):
            q = q.filter(Prestamo.fecha_solicitud >= _parse_fecha(fecha_inicio))
        if _parse_fecha(fecha_fin):
            q = q.filter(Prestamo.fecha_solicitud <= _parse_fecha(fecha_fin) + timedelta(days=1))
        prestamos_equipos = q.order_by(Prestamo.fecha_solicitud.desc()).all()

    if _puede_ver_libros() and tipo_recurso != 'equipos':
        q = PrestamoLibro.query.join(Libro).join(Usuario, PrestamoLibro.id_usuario == Usuario.id_usuario)
        if estado:
            q = q.filter(PrestamoLibro.estado == estado)
        if _parse_fecha(fecha_inicio):
            q = q.filter(PrestamoLibro.fecha_solicitud >= _parse_fecha(fecha_inicio))
        if _parse_fecha(fecha_fin):
            q = q.filter(PrestamoLibro.fecha_solicitud <= _parse_fecha(fecha_fin) + timedelta(days=1))
        prestamos_libros = q.order_by(PrestamoLibro.fecha_solicitud.desc()).all()

    return render_template('reportes/prestamos.html',
        prestamos_equipos=prestamos_equipos,
        prestamos_libros=prestamos_libros,
        estado=estado, tipo_recurso=tipo_recurso,
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin
    )


# ─── MIS PRÉSTAMOS ───────────────────────────────────────────
@bp.route('/mis-prestamos')
@login_required
def mis_prestamos():
    estado = request.args.get('estado', '')

    mis_equipos = Prestamo.query.filter_by(id_usuario=current_user.id_usuario)
    mis_libros = PrestamoLibro.query.filter_by(id_usuario=current_user.id_usuario)

    if estado:
        mis_equipos = mis_equipos.filter(Prestamo.estado == estado)
        mis_libros = mis_libros.filter(PrestamoLibro.estado == estado)

    mis_equipos = mis_equipos.order_by(Prestamo.fecha_solicitud.desc()).all()
    mis_libros = mis_libros.order_by(PrestamoLibro.fecha_solicitud.desc()).all()

    return render_template('reportes/mis_prestamos.html',
        mis_equipos=mis_equipos, mis_libros=mis_libros,
        estado=estado
    )


# ─── USUARIOS ACTIVOS ────────────────────────────────────────
@bp.route('/usuarios-activos')
@login_required
def usuarios_activos():
    if current_user.rol != 'administrador':
        flash('Solo el administrador puede ver este reporte.', 'danger')
        return redirect(url_for('reportes.index'))

    usuarios = Usuario.query.filter_by(estado='activo').all()
    # Conteo de préstamos activos
    usuarios_con_prestamos = []
    for u in usuarios:
        count = u.prestamos_activos_count()
        usuarios_con_prestamos.append({'usuario': u, 'activos': count})
    
    # Ordenar por cantidad de préstamos descendente
    usuarios_con_prestamos.sort(key=lambda x: x['activos'], reverse=True)

    return render_template('reportes/usuarios_activos.html',
        usuarios=usuarios_con_prestamos
    )


# ─── EXPORTACIÓN A EXCEL ─────────────────────────────────────
@bp.route('/exportar/excel/<tipo_reporte>')
@login_required
def exportar_excel(tipo_reporte):
    if _es_usuario_basico() and tipo_reporte != 'mis_prestamos':
        abort(403)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        flash('Librería openpyxl no disponible. Instala los requerimientos.', 'danger')
        return redirect(url_for('reportes.index'))

    wb = Workbook()
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    fecha_hoy = datetime.now().strftime('%Y-%m-%d')

    if tipo_reporte == 'inventario_equipos':
        if not _puede_ver_equipos(): abort(403)
        ws.title = "Inventario Equipos"
        headers = ['ID', 'Nombre', 'Tipo', 'Marca', 'Modelo', 'N° Serie', 'Estado', 'Ubicación', 'Registro']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal='center')
        for row, e in enumerate(Equipo.query.filter_by(eliminado=False).all(), 2):
            ws.cell(row=row, column=1, value=e.id_equipo)
            ws.cell(row=row, column=2, value=e.nombre)
            ws.cell(row=row, column=3, value=e.tipo_equipo)
            ws.cell(row=row, column=4, value=e.marca or '')
            ws.cell(row=row, column=5, value=e.modelo or '')
            ws.cell(row=row, column=6, value=e.numero_serie)
            ws.cell(row=row, column=7, value=e.estado)
            ws.cell(row=row, column=8, value=e.ubicacion or '')
            ws.cell(row=row, column=9, value=str(e.fecha_registro.strftime('%Y-%m-%d') if e.fecha_registro else ''))

    elif tipo_reporte == 'inventario_libros':
        if not _puede_ver_libros(): abort(403)
        ws.title = "Inventario Libros"
        headers = ['ID', 'Título', 'Autor', 'Género', 'Código', 'Estado', 'Ubicación', 'Registro']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal='center')
        for row, l in enumerate(Libro.query.filter_by(eliminado=False).all(), 2):
            ws.cell(row=row, column=1, value=l.id_libro)
            ws.cell(row=row, column=2, value=l.titulo)
            ws.cell(row=row, column=3, value=l.autor)
            ws.cell(row=row, column=4, value=l.genero)
            ws.cell(row=row, column=5, value=l.codigo_unico)
            ws.cell(row=row, column=6, value=l.estado)
            ws.cell(row=row, column=7, value=l.ubicacion or '')
            ws.cell(row=row, column=8, value=str(l.fecha_registro.strftime('%Y-%m-%d') if l.fecha_registro else ''))

    elif tipo_reporte == 'prestamos_equipos':
        if not _puede_ver_equipos(): abort(403)
        ws.title = "Préstamos Equipos"
        headers = ['Cod', 'Usuario', 'Equipo', 'Estado', 'Solicitud', 'Devolución']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal='center')
        for row, p in enumerate(Prestamo.query.order_by(Prestamo.fecha_solicitud.desc()).all(), 2):
            ws.cell(row=row, column=1, value=p.id_prestamo)
            ws.cell(row=row, column=2, value=p.usuario.nombre_completo() if p.usuario else '')
            ws.cell(row=row, column=3, value=p.equipo.nombre if p.equipo else '')
            ws.cell(row=row, column=4, value=p.estado)
            ws.cell(row=row, column=5, value=str(p.fecha_solicitud.strftime('%Y-%m-%d') if p.fecha_solicitud else ''))
            ws.cell(row=row, column=6, value=str(p.fecha_devolucion_real.strftime('%Y-%m-%d') if p.fecha_devolucion_real else 'Pendiente'))

    elif tipo_reporte == 'prestamos_libros':
        if not _puede_ver_libros(): abort(403)
        ws.title = "Préstamos Libros"
        headers = ['Cod', 'Usuario', 'Libro', 'Estado', 'Solicitud', 'Devolución']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal='center')
        for row, p in enumerate(PrestamoLibro.query.order_by(PrestamoLibro.fecha_solicitud.desc()).all(), 2):
            ws.cell(row=row, column=1, value=p.id_prestamo_libro)
            ws.cell(row=row, column=2, value=p.usuario.nombre_completo() if p.usuario else '')
            ws.cell(row=row, column=3, value=p.libro.titulo if p.libro else '')
            ws.cell(row=row, column=4, value=p.estado)
            ws.cell(row=row, column=5, value=str(p.fecha_solicitud.strftime('%Y-%m-%d') if p.fecha_solicitud else ''))
            ws.cell(row=row, column=6, value=str(p.fecha_devolucion_real.strftime('%Y-%m-%d') if p.fecha_devolucion_real else 'Pendiente'))

    elif tipo_reporte == 'mis_prestamos':
        ws.title = "Mis Préstamos"
        headers = ['Tipo', 'Recurso', 'Estado', 'Solicitud', 'Devolución']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal='center')
        row = 2
        for p in Prestamo.query.filter_by(id_usuario=current_user.id_usuario).order_by(Prestamo.fecha_solicitud.desc()).all():
            ws.cell(row=row, column=1, value='Equipo')
            ws.cell(row=row, column=2, value=p.equipo.nombre if p.equipo else '')
            ws.cell(row=row, column=3, value=p.estado)
            ws.cell(row=row, column=4, value=str(p.fecha_solicitud.strftime('%Y-%m-%d') if p.fecha_solicitud else ''))
            ws.cell(row=row, column=5, value=str(p.fecha_devolucion_real.strftime('%Y-%m-%d') if p.fecha_devolucion_real else 'Pendiente'))
            row += 1
        for p in PrestamoLibro.query.filter_by(id_usuario=current_user.id_usuario).order_by(PrestamoLibro.fecha_solicitud.desc()).all():
            ws.cell(row=row, column=1, value='Libro')
            ws.cell(row=row, column=2, value=p.libro.titulo if p.libro else '')
            ws.cell(row=row, column=3, value=p.estado)
            ws.cell(row=row, column=4, value=str(p.fecha_solicitud.strftime('%Y-%m-%d') if p.fecha_solicitud else ''))
            ws.cell(row=row, column=5, value=str(p.fecha_devolucion_real.strftime('%Y-%m-%d') if p.fecha_devolucion_real else 'Pendiente'))
            row += 1
    else:
        flash('Reporte no válido.', 'warning')
        return redirect(url_for('reportes.index'))

    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(max_length + 2, 10)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"reporte_{tipo_reporte}_{fecha_hoy}.xlsx"

    current_app.logger.info(f"Exportación Excel '{tipo_reporte}' por usuario ID {current_user.id_usuario}.")
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── EXPORTACIÓN A PDF ───────────────────────────────────────
@bp.route('/exportar/pdf/<tipo_reporte>')
@login_required
def exportar_pdf(tipo_reporte):
    if _es_usuario_basico() and tipo_reporte != 'mis_prestamos':
        abort(403)

    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        flash('Librería reportlab no disponible. Instala los requerimientos.', 'danger')
        return redirect(url_for('reportes.index'))

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []
    fecha_hoy = datetime.now().strftime('%Y-%m-%d %H:%M')

    title_style = styles['Title']
    subtitle_style = styles['Normal']

    header_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39A900')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
    ])

    if tipo_reporte == 'inventario_equipos':
        if not _puede_ver_equipos(): abort(403)
        elements.append(Paragraph('Reporte de Inventario - Equipos', title_style))
        elements.append(Paragraph(f'Generado: {fecha_hoy} | Por: {current_user.nombre_completo()}', subtitle_style))
        elements.append(Spacer(1, 20))
        data = [['ID', 'Nombre', 'Tipo', 'Marca', 'N° Serie', 'Estado', 'Ubicación']]
        for e in Equipo.query.filter_by(eliminado=False).all():
            data.append([e.id_equipo, e.nombre[:25], e.tipo_equipo, e.marca or '', e.numero_serie, e.estado, e.ubicacion or ''])
        t = Table(data, repeatRows=1)
        t.setStyle(header_style)
        elements.append(t)

    elif tipo_reporte == 'inventario_libros':
        if not _puede_ver_libros(): abort(403)
        elements.append(Paragraph('Reporte de Inventario - Libros', title_style))
        elements.append(Paragraph(f'Generado: {fecha_hoy} | Por: {current_user.nombre_completo()}', subtitle_style))
        elements.append(Spacer(1, 20))
        data = [['ID', 'Título', 'Autor', 'Género', 'Código', 'Estado', 'Ubicación']]
        for l in Libro.query.filter_by(eliminado=False).all():
            data.append([l.id_libro, l.titulo[:30], l.autor[:20], l.genero, l.codigo_unico, l.estado, l.ubicacion or ''])
        t = Table(data, repeatRows=1)
        t.setStyle(header_style)
        elements.append(t)

    elif tipo_reporte == 'prestamos_equipos':
        if not _puede_ver_equipos(): abort(403)
        elements.append(Paragraph('Reporte de Préstamos - Equipos', title_style))
        elements.append(Paragraph(f'Generado: {fecha_hoy}', subtitle_style))
        elements.append(Spacer(1, 20))
        data = [['Cod', 'Usuario', 'Equipo', 'Estado', 'Solicitud', 'Devolución']]
        for p in Prestamo.query.order_by(Prestamo.fecha_solicitud.desc()).all():
            data.append([
                p.id_prestamo, p.usuario.nombre_completo()[:25] if p.usuario else '',
                p.equipo.nombre[:20] if p.equipo else '', p.estado,
                p.fecha_solicitud.strftime('%Y-%m-%d') if p.fecha_solicitud else '',
                p.fecha_devolucion_real.strftime('%Y-%m-%d') if p.fecha_devolucion_real else 'Pendiente'
            ])
        t = Table(data, repeatRows=1)
        t.setStyle(header_style)
        elements.append(t)

    elif tipo_reporte == 'prestamos_libros':
        if not _puede_ver_libros(): abort(403)
        elements.append(Paragraph('Reporte de Préstamos - Libros', title_style))
        elements.append(Paragraph(f'Generado: {fecha_hoy}', subtitle_style))
        elements.append(Spacer(1, 20))
        data = [['Cod', 'Usuario', 'Libro', 'Estado', 'Solicitud', 'Devolución']]
        for p in PrestamoLibro.query.order_by(PrestamoLibro.fecha_solicitud.desc()).all():
            data.append([
                p.id_prestamo_libro, p.usuario.nombre_completo()[:25] if p.usuario else '',
                p.libro.titulo[:25] if p.libro else '', p.estado,
                p.fecha_solicitud.strftime('%Y-%m-%d') if p.fecha_solicitud else '',
                p.fecha_devolucion_real.strftime('%Y-%m-%d') if p.fecha_devolucion_real else 'Pendiente'
            ])
        t = Table(data, repeatRows=1)
        t.setStyle(header_style)
        elements.append(t)

    elif tipo_reporte == 'mis_prestamos':
        elements.append(Paragraph('Mi Historial de Préstamos', title_style))
        elements.append(Paragraph(f'{current_user.nombre_completo()} | {fecha_hoy}', subtitle_style))
        elements.append(Spacer(1, 20))
        data = [['Tipo', 'Recurso', 'Estado', 'Solicitud', 'Devolución']]
        for p in Prestamo.query.filter_by(id_usuario=current_user.id_usuario).order_by(Prestamo.fecha_solicitud.desc()).all():
            data.append([
                'Equipo', p.equipo.nombre[:25] if p.equipo else '', p.estado,
                p.fecha_solicitud.strftime('%Y-%m-%d') if p.fecha_solicitud else '',
                p.fecha_devolucion_real.strftime('%Y-%m-%d') if p.fecha_devolucion_real else 'Pendiente'
            ])
        for p in PrestamoLibro.query.filter_by(id_usuario=current_user.id_usuario).order_by(PrestamoLibro.fecha_solicitud.desc()).all():
            data.append([
                'Libro', p.libro.titulo[:25] if p.libro else '', p.estado,
                p.fecha_solicitud.strftime('%Y-%m-%d') if p.fecha_solicitud else '',
                p.fecha_devolucion_real.strftime('%Y-%m-%d') if p.fecha_devolucion_real else 'Pendiente'
            ])
        t = Table(data, repeatRows=1)
        t.setStyle(header_style)
        elements.append(t)
    else:
        flash('Reporte no válido.', 'warning')
        return redirect(url_for('reportes.index'))

    if len(elements) < 3:
        flash('No hay datos para exportar.', 'warning')
        return redirect(url_for('reportes.index'))

    doc.build(elements)
    output.seek(0)
    filename = f"reporte_{tipo_reporte}_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    
    current_app.logger.info(f"Exportación PDF '{tipo_reporte}' por usuario ID {current_user.id_usuario}.")
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/pdf')
